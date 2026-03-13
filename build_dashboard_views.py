"""
Build Variance sheets (Individual, By Department, By Country, By Group) and Hours view.
- Variance: Annual Leave (Allowance, Taken, Remaining) and WFH (Allowed from first BLIP date to today, 1/week, Taken, Variance). Roll-up by Department, Country, Group.
- Hours view: one row per date; per employee In, Out, Break, Var; Total Var row.

Usage:
  python build_dashboard_views.py [--input PATH] [--output PATH] [--expected-hours 7.5] [--wfh-allowance OVERRIDE]

WFH allowed = weeks from first BLIP entry date to today (1 per week). Override with --wfh-allowance if needed. Close workbook before running if updating in place.
"""
import os
import math
from datetime import date
import argparse
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font

_PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INPUT = os.path.join(_PROJECT_ROOT, "Variance_Excel.xlsx")
EXPECTED_HOURS_PER_DAY = 7.5
WFH_ALLOWANCE_DEFAULT = None  # None = compute from BLIP (weeks from first entry to today)
# Per-person WFH allowance overrides: working remotely (0), special arrangements (2 or 3), else default 9
WFH_ALLOWANCE_OVERRIDES = {
    "Albano Limas": 0,
    "John Hetherton": 0,
    "Elizabeth Kinnear-Mellor": 0,
    "Jamie Rixton": 0,
    "Ryan Holland": 0,
    "Otto Carlisle": 2,
    "Mark Turner": 3,
    "Elias": 3,
    "Fabian": 3,
    "William Betts": 3,
}
WFH_ALLOWANCE_DEFAULT_WHEN_OVERRIDE = 9
# Period is Jan 1 to date (~9 weeks). Override 2 = 2/week → 18, 3 = 3/week → 27. 0 and 9 (1/week) stay as-is.
WFH_WEEKS = 9

# Fully remote employees (0 WFH allowance) — exclude from all BLIP / Time sheets
FULLY_REMOTE_EMPLOYEES = frozenset(
    k.lower() for k, v in WFH_ALLOWANCE_OVERRIDES.items() if v == 0
)

# Formatting: 4-step gradient (positive=green, negative=red, zero=neutral gray)
# Better contrast and smoother magnitude mapping
FILL_GREEN_VLIGHT = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")   # 0–0.5
FILL_GREEN_LIGHT = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")    # 0.5–2
FILL_GREEN_MED = PatternFill(start_color="81C784", end_color="81C784", fill_type="solid")    # 2–5
FILL_GREEN_DARK = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")   # 5+
FILL_RED_VLIGHT = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")   # 0–0.5
FILL_RED_LIGHT = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")   # 0.5–2
FILL_RED_MED = PatternFill(start_color="E57373", end_color="E57373", fill_type="solid")      # 2–5
FILL_RED_DARK = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")    # 5+
FILL_NEUTRAL = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")     # zero

# Gradient thresholds (absolute value)
VARIANCE_T1, VARIANCE_T2, VARIANCE_T3 = 0.5, 2.0, 5.0


def _first_last(s):
    """Reduce to First + Last name for matching (align with build_dashboard_excel)."""
    parts = str(s).strip().split() if s else []
    if len(parts) >= 2:
        return parts[0] + " " + parts[-1]
    return parts[0] if parts else ""


def _uk_holidays_for_year(year):
    """UK England & Wales bank holidays for the given year. Returns set of pd.Timestamp dates (normalized)."""
    # Fixed and variable dates for 2026, 2027 (extend as needed)
    holidays = set()
    # New Year
    holidays.add(pd.Timestamp(year, 1, 1))
    # Early May (first Monday May)
    d = pd.Timestamp(year, 5, 1)
    while d.weekday() != 0:
        d += pd.Timedelta(days=1)
    holidays.add(d)
    # Spring (last Monday May)
    d = pd.Timestamp(year, 5, 31)
    while d.weekday() != 0:
        d -= pd.Timedelta(days=1)
    holidays.add(d)
    # Summer (last Monday August)
    d = pd.Timestamp(year, 8, 31)
    while d.weekday() != 0:
        d -= pd.Timedelta(days=1)
    holidays.add(d)
    # Christmas / Boxing Day
    holidays.add(pd.Timestamp(year, 12, 25))
    holidays.add(pd.Timestamp(year, 12, 26))
    # Good Friday / Easter Monday (simplified: use datetime or hardcode for 2026/2027)
    if year == 2026:
        holidays.add(pd.Timestamp(2026, 4, 3))
        holidays.add(pd.Timestamp(2026, 4, 6))
    elif year == 2027:
        holidays.add(pd.Timestamp(2027, 3, 26))
        holidays.add(pd.Timestamp(2027, 3, 29))
    return {pd.Timestamp(d).normalize() for d in holidays}


def _get_time_sheet_date_range(df_blip, df_absence, time_from_arg=None, time_to_arg=None):
    """Return (start, end) as pd.Timestamp for the Time by Employee continuous calendar. Prefer args; else derive from BLIP then Absence."""
    if time_from_arg and time_to_arg:
        start = pd.Timestamp(time_from_arg).normalize()
        end = pd.Timestamp(time_to_arg).normalize()
        if start > end:
            start, end = end, start
        return start, end
    start, end = None, None
    if not df_blip.empty:
        date_col = None
        if "date" in df_blip.columns:
            date_col = df_blip["date"]
        elif "Clock In Date" in df_blip.columns:
            date_col = df_blip["Clock In Date"]
        if date_col is not None:
            dt = pd.to_datetime(date_col, errors="coerce").dropna()
            if not dt.empty:
                start = pd.Timestamp(dt.min()).normalize()
                end = pd.Timestamp(dt.max()).normalize()
    if (start is None or end is None) and not df_absence.empty:
        for col in ["start_dt", "Absence start date"]:
            if col in df_absence.columns:
                s = pd.to_datetime(df_absence[col], errors="coerce").dropna()
                if not s.empty and start is None:
                    start = pd.Timestamp(s.min()).normalize()
                break
        for col in ["end_dt", "Absence end date"]:
            if col in df_absence.columns:
                e = pd.to_datetime(df_absence[col], errors="coerce").dropna()
                if not e.empty and end is None:
                    end = pd.Timestamp(e.max()).normalize()
                break
    if start is None or end is None:
        # Default: current month
        today = pd.Timestamp.now().normalize()
        start = today.replace(day=1)
        end = today
    if start > end:
        end = start
    return start, end


def _build_absence_daily_lookup(df_absence):
    """Build (employee_short, date) -> set of absence_category. Uses First+Last for employee. Returns dict."""
    if df_absence is None or df_absence.empty:
        return {}
    df = df_absence.copy()
    if "employee" not in df.columns:
        return {}
    df["employee"] = df["employee"].fillna("").astype(str).apply(_normalize_employee).apply(_first_last)
    start_col = next((c for c in ["start_dt", "Absence start date"] if c in df.columns), None)
    end_col = next((c for c in ["end_dt", "Absence end date"] if c in df.columns), None)
    if not start_col:
        return {}
    df["start_dt"] = pd.to_datetime(df[start_col], errors="coerce")
    df["end_dt"] = pd.to_datetime(df[end_col], errors="coerce") if end_col else df["start_dt"]
    df["end_dt"] = df["end_dt"].fillna(df["start_dt"])
    df.loc[df["end_dt"] < df["start_dt"], "end_dt"] = df["start_dt"]
    cat_col = next((c for c in ["absence_category", "Absence type"] if c in df.columns), None)
    if not cat_col:
        return {}
    try:
        from absence_daily import expand_to_daily
        expanded = expand_to_daily(df[["employee", "start_dt", "end_dt", cat_col]].copy())
    except Exception:
        return {}
    if expanded.empty:
        return {}
    expanded["date"] = pd.to_datetime(expanded["date"], errors="coerce").dt.normalize()
    lookup = {}
    def _norm_cat(raw_val):
        c = (str(raw_val) or "").strip().lower()
        if "medical" in c or "sick" in c:
            return "Medical + Sickness"
        if "annual" in c or "holiday" in c:
            return "Annual Leave"
        return None

    for _, r in expanded.iterrows():
        emp = r["employee"].strip()
        dt = r["date"]
        if pd.isna(dt) or not emp:
            continue
        key = (emp, pd.Timestamp(dt).normalize())
        raw = str(r.get(cat_col, "") or "").strip()
        cat = _norm_cat(raw) if raw else None
        if key not in lookup:
            lookup[key] = set()
        if cat:
            lookup[key].add(cat)
    return lookup


def _working_days_in_range(start_ts, end_ts, national_holidays):
    """Count weekdays (Mon=0 .. Fri=4) in [start_ts, end_ts] excluding national_holidays."""
    n = 0
    d = pd.Timestamp(start_ts).normalize()
    end = pd.Timestamp(end_ts).normalize()
    while d <= end:
        if d.weekday() < 5 and d not in national_holidays:
            n += 1
        d += pd.Timedelta(days=1)
    return n


def _build_time_summary_rows(employees_list, time_sheet_dates, hours_lookup, absence_lookup, national_holidays, expected_hours_per_day=EXPECTED_HOURS_PER_DAY):
    """
    Build rows for Time Summary sheet: Employee, Current Week hours, %, MTD, %, YTD, %.
    Uses same var logic as Time by Employee (weekend/holiday exclude, sick exclude, annual leave=0).
    """
    if not time_sheet_dates or not employees_list:
        return []
    time_sheet_end = pd.Timestamp(time_sheet_dates[-1]).normalize()

    def var_value(emp, dt):
        ts = pd.Timestamp(dt).normalize()
        if ts.weekday() >= 5 or ts in national_holidays:
            return None
        cats = absence_lookup.get((emp, ts), set())
        if "Medical + Sickness" in cats:
            return None
        if "Annual Leave" in cats:
            return 0.0
        return float(hours_lookup.get((emp, ts), {}).get("var_hrs", 0) or 0)

    # Current week (Mon-Sun containing time_sheet_end)
    week_start = time_sheet_end - pd.Timedelta(days=time_sheet_end.weekday())
    week_end = week_start + pd.Timedelta(days=6)
    dates_in_week = [d for d in time_sheet_dates if week_start <= pd.Timestamp(d).normalize() <= week_end]
    working_days_week = _working_days_in_range(week_start, week_end, national_holidays)
    expected_week = working_days_week * expected_hours_per_day if working_days_week else None

    # MTD: first of month to time_sheet_end
    month_start = time_sheet_end.replace(day=1)
    dates_mtd = [d for d in time_sheet_dates if month_start <= pd.Timestamp(d).normalize() <= time_sheet_end]
    working_days_mtd = _working_days_in_range(month_start, time_sheet_end, national_holidays)
    expected_mtd = working_days_mtd * expected_hours_per_day if working_days_mtd else None

    # YTD: Jan 1 to time_sheet_end
    year_start = time_sheet_end.replace(month=1, day=1)
    dates_ytd = [d for d in time_sheet_dates if year_start <= pd.Timestamp(d).normalize() <= time_sheet_end]
    working_days_ytd = _working_days_in_range(year_start, time_sheet_end, national_holidays)
    expected_ytd = working_days_ytd * expected_hours_per_day if working_days_ytd else None

    rows = []
    for emp in employees_list:
        cw_sum = sum(v for d in dates_in_week for v in [var_value(emp, d)] if v is not None)
        mtd_sum = sum(v for d in dates_mtd for v in [var_value(emp, d)] if v is not None)
        ytd_sum = sum(v for d in dates_ytd for v in [var_value(emp, d)] if v is not None)
        cw_pct = round((cw_sum / expected_week * 100)) if expected_week and expected_week != 0 else None
        mtd_pct = round((mtd_sum / expected_mtd * 100)) if expected_mtd and expected_mtd != 0 else None
        ytd_pct = round((ytd_sum / expected_ytd * 100)) if expected_ytd and expected_ytd != 0 else None
        rows.append({
            "Employee": emp,
            "Current Week hours": round(cw_sum, 2),
            "Current Week %": cw_pct,
            "MTD": round(mtd_sum, 2),
            "MTD %": mtd_pct,
            "YTD": round(ytd_sum, 2),
            "YTD %": ytd_pct,
        })
    return rows


def _round_to_half(x):
    """Round to nearest 0.5 (whole or half only)."""
    try:
        v = float(x)
    except (TypeError, ValueError):
        return x
    return round(v * 2) / 2


def _break_hrs_to_hhmm(break_hrs):
    """Format decimal hours as HH:MM (e.g. 0.9 -> 00:54)."""
    total_mins = int(round((break_hrs or 0) * 60))
    h = total_mins // 60
    m = total_mins % 60
    return f"{h:02d}:{m:02d}"


def var_hrs_to_hours_str(v):
    """Format variance hours as 'X.Xh' or '-X.Xh' for display."""
    try:
        x = float(v)
    except (TypeError, ValueError):
        return "" if v is None or (isinstance(v, str) and str(v).strip() == "") else str(v)
    return f"{x:.1f}h" if abs(x - round(x)) > 1e-6 else f"{int(round(x))}h"


def _fill_for_variance(val):
    """Return PatternFill for a variance value: 4-step green/red gradient, neutral for zero."""
    try:
        v = float(val)
    except (TypeError, ValueError):
        return None
    av = abs(v)
    if v > 0:
        if av <= VARIANCE_T1:
            return FILL_GREEN_VLIGHT
        if av <= VARIANCE_T2:
            return FILL_GREEN_LIGHT
        if av <= VARIANCE_T3:
            return FILL_GREEN_MED
        return FILL_GREEN_DARK
    if v < 0:
        if av <= VARIANCE_T1:
            return FILL_RED_VLIGHT
        if av <= VARIANCE_T2:
            return FILL_RED_LIGHT
        if av <= VARIANCE_T3:
            return FILL_RED_MED
        return FILL_RED_DARK
    return FILL_NEUTRAL


# CSS equivalents for Streamlit (4-step gradient)
_VARIANCE_CSS = {
    "green_vlight": "background-color: #E8F5E9",
    "green_light": "background-color: #C8E6C9",
    "green_med": "background-color: #81C784; color: white",
    "green_dark": "background-color: #2E7D32; color: white",
    "red_vlight": "background-color: #FFEBEE",
    "red_light": "background-color: #FFCDD2",
    "red_med": "background-color: #E57373; color: white",
    "red_dark": "background-color: #C62828; color: white",
    "neutral": "background-color: #FAFAFA",
}


def variance_bg_color(val):
    """Return CSS background color for a variance value (for Streamlit/pandas Styler)."""
    try:
        v = float(val)
    except (TypeError, ValueError):
        return ""
    av = abs(v)
    if v > 0:
        if av <= VARIANCE_T1:
            return _VARIANCE_CSS["green_vlight"]
        if av <= VARIANCE_T2:
            return _VARIANCE_CSS["green_light"]
        if av <= VARIANCE_T3:
            return _VARIANCE_CSS["green_med"]
        return _VARIANCE_CSS["green_dark"]
    if v < 0:
        if av <= VARIANCE_T1:
            return _VARIANCE_CSS["red_vlight"]
        if av <= VARIANCE_T2:
            return _VARIANCE_CSS["red_light"]
        if av <= VARIANCE_T3:
            return _VARIANCE_CSS["red_med"]
        return _VARIANCE_CSS["red_dark"]
    return _VARIANCE_CSS["neutral"]


def _normalize_employee(s):
    return " ".join(str(s).strip().split()) if s else ""


def wfh_allowed_from_blip(df_blip):
    """WFH allowed = 1 per week from first BLIP entry date to today. Returns int (weeks)."""
    if df_blip is None or df_blip.empty:
        return 0
    date_col = next((c for c in df_blip.columns if "date" in str(c).lower() or "Date" in str(c)), None)
    if not date_col:
        return 0
    dates = pd.to_datetime(df_blip[date_col], errors="coerce", dayfirst=True).dropna()
    if dates.empty:
        return 0
    first_dt = dates.min()
    if pd.isna(first_dt):
        return 0
    first_date = first_dt.date() if hasattr(first_dt, "date") else first_dt
    today = date.today()
    days = (today - first_date).days
    return max(0, math.ceil(days / 7.0))


def wfh_allowed_per_employee_from_first_absence(df_absence, df_blip):
    """
    WFH allowed per employee = 1 per week from that employee's first absence entry (min start_dt) to today.
    If employee has no absence, use that employee's first BLIP date; if no BLIP either, use global first BLIP date.
    Returns Series index=normalized employee name, value=WFH allowed (int).
    """
    today = date.today()
    out = {}

    # First absence date per employee (from Absence data)
    if df_absence is not None and not df_absence.empty and "employee" in df_absence.columns:
        start_col = next((c for c in df_absence.columns if ("start" in str(c).lower() and "date" in str(c).lower()) or str(c).strip() == "start_dt"), None)
        if start_col:
            ab = df_absence.copy()
            ab["_emp"] = ab["employee"].astype(str).apply(_normalize_employee)
            ab["_start"] = pd.to_datetime(ab[start_col], errors="coerce", dayfirst=True)
            first_abs = ab.dropna(subset=["_start"]).groupby("_emp")["_start"].min()
            for emp, first_dt in first_abs.items():
                first_date = first_dt.date() if hasattr(first_dt, "date") else first_dt
                days = (today - first_date).days
                out[emp] = max(0, math.ceil(days / 7.0))

    # First BLIP date per employee (for fallback)
    blip_first_global = None
    blip_first_per_emp = {}
    if df_blip is not None and not df_blip.empty:
        date_col = next((c for c in df_blip.columns if "date" in str(c).lower() or "Date" in str(c)), None)
        if date_col:
            blip = df_blip.copy()
            blip["_emp"] = blip["employee"].fillna("").astype(str).apply(_normalize_employee)
            blip["_date"] = pd.to_datetime(blip[date_col], errors="coerce", dayfirst=True)
            blip = blip.dropna(subset=["_date"])
            if not blip.empty:
                blip_first_global = blip["_date"].min()
                blip_first_global = blip_first_global.date() if hasattr(blip_first_global, "date") else blip_first_global
                first_per_emp = blip.groupby("_emp")["_date"].min()
                for emp, first_dt in first_per_emp.items():
                    first_date = first_dt.date() if hasattr(first_dt, "date") else first_dt
                    days = (today - first_date).days
                    blip_first_per_emp[emp] = max(0, math.ceil(days / 7.0))

    # Fill missing: use employee's first BLIP, else default 4 (no absence = 4 WFH days)
    global_fallback = 4
    return out, blip_first_per_emp, global_fallback


def _wfh_allowance_for_employee(emp, default_val):
    """Look up per-person WFH override. Case-insensitive exact match; single-word keys also prefix-match (e.g. 'Elias' matches 'Elias Zimmerman')."""
    emp_norm = (_normalize_employee(emp) if isinstance(emp, str) else "") or ""
    emp_lower = emp_norm.lower()
    for key, val in WFH_ALLOWANCE_OVERRIDES.items():
        key_norm = _normalize_employee(key)
        key_lower = key_norm.lower()
        if emp_lower == key_lower:
            return val
        # Single-word keys (first name only) → prefix match
        if " " not in key_norm and emp_lower.startswith(key_lower + " "):
            return val
    return default_val


def build_variance_individual_df(df_employees, df_blip, df_absence=None, wfh_allowance_override=None):
    """
    Build individual variance table: Employee, Team, Country, Group, Annual (Allowance, Taken, % YTD, Remaining),
    WFH (Allowed, Taken, Variance), Medical/External/Other Taken. WFH Allowed uses per-person overrides then default.
    """
    if df_employees.empty:
        return pd.DataFrame()

    df = df_employees.copy()
    allowance_col = next((c for c in df.columns if "Leave Allowance" in str(c) or c == "Leave Allowance"), None)
    annual_col = next((c for c in df.columns if str(c).strip() == "Annual Leave"), None)
    wfh_col = next((c for c in df.columns if str(c).strip() == "WFH"), None)
    medical_col = next((c for c in df.columns if "Medical" in str(c)), None)
    external_col = next((c for c in df.columns if "External" in str(c)), None)
    other_col = next((c for c in df.columns if "Other" in str(c) and "WFH" not in str(c)), None)
    team_col = next((c for c in df.columns if "Team" in str(c)), None)
    country_col = next((c for c in df.columns if "Country" in str(c)), None)

    out = pd.DataFrame()
    out["Employee"] = df["Employee"].astype(str).apply(_normalize_employee)
    out["Team"] = df[team_col].fillna("").astype(str) if team_col else ""
    out["Country"] = df[country_col].fillna("").astype(str) if country_col else ""

    from absence_daily import infer_organisation_from_team
    out["Group"] = infer_organisation_from_team(out["Team"]).values

    allowance = pd.to_numeric(df[allowance_col], errors="coerce").fillna(0) if allowance_col else pd.Series(0, index=df.index)
    annual_taken = pd.to_numeric(df[annual_col], errors="coerce").fillna(0) if annual_col else pd.Series(0, index=df.index)
    out["Annual Leave Allowance"] = allowance.apply(_round_to_half)
    out["Annual Leave Taken"] = annual_taken.apply(_round_to_half)
    out["% YTD"] = (out["Annual Leave Taken"] / out["Annual Leave Allowance"].replace(0, float("nan")) * 100).round(1)
    out["Annual Leave Remaining"] = (allowance - annual_taken).apply(_round_to_half)

    if wfh_allowance_override is not None:
        default_wfh = WFH_ALLOWANCE_DEFAULT_WHEN_OVERRIDE
    else:
        first_abs_map, blip_per_emp, default_no_absence = wfh_allowed_per_employee_from_first_absence(df_absence, df_blip)
        default_wfh = default_no_absence if default_no_absence is not None else 4
    wfh_list = [_wfh_allowance_for_employee(emp, default_wfh) for emp in out["Employee"]]
    # Convert per-week to total: 2/week → 18, 3/week → 27; 0 and 9 (1/week total) unchanged
    out["WFH Allowance"] = [
        v * WFH_WEEKS if v in (2, 3) else v for v in wfh_list
    ]

    wfh_taken = pd.to_numeric(df[wfh_col], errors="coerce").fillna(0) if wfh_col else pd.Series(0.0, index=df.index)
    out["WFH Taken"] = wfh_taken.apply(_round_to_half)
    out["WFH Variance"] = (out["WFH Allowance"] - out["WFH Taken"]).apply(_round_to_half)

    med = pd.to_numeric(df[medical_col], errors="coerce").fillna(0) if medical_col else pd.Series(0.0, index=df.index)
    ext = pd.to_numeric(df[external_col], errors="coerce").fillna(0) if external_col else pd.Series(0.0, index=df.index)
    oth = pd.to_numeric(df[other_col], errors="coerce").fillna(0) if other_col else pd.Series(0.0, index=df.index)
    out["Medical/Sick Taken"] = med.apply(_round_to_half)
    out["External Assignments Taken"] = ext.apply(_round_to_half)
    out["Other Taken"] = oth.apply(_round_to_half)

    return out


def build_hours_view_data(df_blip, expected_hours=EXPECTED_HOURS_PER_DAY):
    """
    From BLIP sheet: group by employee (First+Last) and date (shift rows only).
    Return (dates_sorted, employees_sorted, lookup (emp_short, date) -> {in, out, break_hrs, var_hrs}).
    """
    if df_blip.empty:
        return [], [], {}

    df = df_blip.copy()
    df["employee"] = df["employee"].fillna("").astype(str).apply(_normalize_employee)
    df["employee_short"] = df["employee"].apply(_first_last)

    # Exclude fully remote employees from Time sheets
    if FULLY_REMOTE_EMPLOYEES:
        _is_remote = df["employee_short"].str.lower().apply(
            lambda x: any(x == r or x.startswith(r + " ") or x.endswith(" " + r) for r in FULLY_REMOTE_EMPLOYEES)
        )
        df = df[~_is_remote]
        if df.empty:
            return [], [], {}

    if "date" not in df.columns and "Clock In Date" in df.columns:
        df["date"] = pd.to_datetime(df["Clock In Date"], errors="coerce")
    else:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["date_norm"] = df["date"].dt.normalize()

    shift_mask = df["blip_type_norm"].astype(str).str.strip().str.lower() == "shift" if "blip_type_norm" in df.columns else pd.Series(True, index=df.index)
    blip = df[shift_mask].copy()

    if blip.empty:
        return [], [], {}

    if "date_norm" not in blip.columns:
        blip["date_norm"] = pd.to_datetime(blip["date"]).dt.normalize()
    daily = blip.groupby(["employee_short", "date_norm"], as_index=False).agg(
        clockin_dt=("clockin_dt", "min"),
        clockout_dt=("clockout_dt", "max"),
        worked_hours=("worked_hours", "sum"),
        duration_hours=("duration_hours", "sum"),
    )
    daily["break_hours"] = (daily["duration_hours"] - daily["worked_hours"]).clip(lower=0)
    daily["var_hours"] = daily["worked_hours"] - expected_hours

    dates_sorted = sorted(daily["date_norm"].dropna().unique())
    employees_sorted = sorted(daily["employee_short"].dropna().unique())
    if not employees_sorted:
        return [], [], {}

    lookup = {}
    for _, r in daily.iterrows():
        emp = r["employee_short"]
        dt = r["date_norm"]
        if pd.isna(dt):
            continue
        key = (emp, pd.Timestamp(dt).normalize())
        in_t = r.get("clockin_dt")
        out_t = r.get("clockout_dt")
        in_str = in_t.strftime("%H:%M") if pd.notna(in_t) and hasattr(in_t, "strftime") else ""
        out_str = out_t.strftime("%H:%M") if pd.notna(out_t) and hasattr(out_t, "strftime") else ""
        break_hrs = float(r.get("break_hours", 0) or 0)
        var_hrs = float(r.get("var_hours", 0) or 0)
        lookup[key] = {"in": in_str, "out": out_str, "break_hrs": break_hrs, "var_hrs": var_hrs}

    return dates_sorted, employees_sorted, lookup


VARIANCE_NUMERIC_COLS = [
    "Annual Leave Allowance", "Annual Leave Taken", "Annual Leave Remaining",
    "WFH Allowance", "WFH Taken", "WFH Variance",
    "Medical/Sick Taken", "External Assignments Taken", "Other Taken",
]
# % YTD is computed, not summed; insert after Annual Leave Taken in output
VARIANCE_NUMERIC_COLS_WITH_PCT_YTD = [
    "Annual Leave Allowance", "Annual Leave Taken", "% YTD", "Annual Leave Remaining",
    "WFH Allowance", "WFH Taken", "WFH Variance",
    "Medical/Sick Taken", "External Assignments Taken", "Other Taken",
]


def build_variance_rollup(df_var, group_col, name_col):
    """One row per group: Name, Headcount, sum of numeric cols; % YTD = (sum Taken / sum Allowance)*100."""
    if df_var.empty:
        return pd.DataFrame()
    cols = [c for c in VARIANCE_NUMERIC_COLS if c in df_var.columns]
    agg = {"Employee": "count"}
    for c in cols:
        agg[c] = "sum"
    out = df_var.groupby(group_col, as_index=False).agg(agg).rename(columns={"Employee": "Headcount", group_col: name_col})
    out["% YTD"] = (out["Annual Leave Taken"] / out["Annual Leave Allowance"].replace(0, float("nan")) * 100).round(1)
    out_cols = [name_col, "Headcount", "Annual Leave Allowance", "Annual Leave Taken", "% YTD", "Annual Leave Remaining"]
    out_cols += [c for c in cols if c not in ("Annual Leave Allowance", "Annual Leave Taken", "Annual Leave Remaining")]
    out = out[[c for c in out_cols if c in out.columns]]
    for c in cols:
        if c in out.columns and c != "% YTD":
            out[c] = out[c].apply(_round_to_half)
    return out


def build_time_rollup(df_employees, hours_lookup, dates_list, employees_list, group_col, name_col):
    """
    One row per group: Name, Headcount, Total Var (sum of per-employee total var_hrs).
    group_col is the column name (Team, Country, or Group); name_col is the label (Department, Country, Group).
    """
    if not employees_list or not dates_list or (df_employees.empty or "Employee" not in df_employees.columns):
        return pd.DataFrame(columns=[name_col, "Headcount", "Total Var"])
    emp_to_total_var = {}
    for emp in employees_list:
        total = 0.0
        for dt in dates_list:
            key = (emp, pd.Timestamp(dt).normalize() if hasattr(dt, "normalize") else dt)
            total += hours_lookup.get(key, {}).get("var_hrs", 0)
        emp_to_total_var[emp] = round(total, 2)
    df_emp = df_employees[["Employee", "Team", "Country"]].copy()
    if group_col == "Group":
        from absence_daily import infer_organisation_from_team
        df_emp["Group"] = infer_organisation_from_team(df_emp["Team"].fillna("")).values
    if group_col not in df_emp.columns:
        return pd.DataFrame(columns=[name_col, "Headcount", "Total Var"])
    df_emp["Total Var"] = df_emp["Employee"].map(emp_to_total_var).fillna(0)
    df_emp = df_emp[df_emp["Employee"].isin(employees_list)]
    if df_emp.empty:
        return pd.DataFrame(columns=[name_col, "Headcount", "Total Var"])
    out = df_emp.groupby(group_col, as_index=False).agg(Employee=("Employee", "count"), Total_Var=("Total Var", "sum"))
    out = out.rename(columns={"Employee": "Headcount", "Total_Var": "Total Var", group_col: name_col})
    out = out[[name_col, "Headcount", "Total Var"]].round(2)
    return out


def _write_variance_sheet(ws, df, title, variance_col_indices):
    """Write dataframe to sheet with title row, headers, data. variance_col_indices = 1-based column indices for variance columns. Use same conditional formatting as time: _fill_for_variance(value) from data."""
    ws["A1"] = title
    headers = list(df.columns)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    for r, row in df.iterrows():
        for c, h in enumerate(headers, start=1):
            val = row.get(h, "")
            cell = ws.cell(row=r + 3, column=c, value=val)
            if c in variance_col_indices:
                fill = _fill_for_variance(val)
                if fill:
                    cell.fill = fill


def _write_time_rollup_sheet(ws, df, title):
    """Write time rollup DataFrame (Name, Headcount, Total Var) with title and conditional formatting on Total Var (same as time Var: _fill_for_variance from data)."""
    ws["A1"] = title
    headers = list(df.columns)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    total_var_col = 3  # 1-based
    for r, row in df.iterrows():
        for c, h in enumerate(headers, start=1):
            val = row.get(h, "")
            cell = ws.cell(row=r + 3, column=c, value=val)
            if c == total_var_col:
                fill = _fill_for_variance(val)
                if fill:
                    cell.fill = fill


def main():
    parser = argparse.ArgumentParser(description="Add Absences view and Hours view sheets to dashboard workbook.")
    parser.add_argument("--input", "-i", default=DEFAULT_INPUT, help="Input workbook (must have Employees and BLIP sheets)")
    parser.add_argument("--output", "-o", help="Output workbook (default: same as input)")
    parser.add_argument("--expected-hours", type=float, default=EXPECTED_HOURS_PER_DAY, help="Expected hours per day for Var")
    parser.add_argument("--wfh-allowance", type=float, default=None, help="Override WFH allowance (default: weeks from first BLIP date to today)")
    parser.add_argument("--time-from", default=None, help="Time by Employee sheet: first date (YYYY-MM-DD). Default: from BLIP/Absence")
    parser.add_argument("--time-to", default=None, help="Time by Employee sheet: last date (YYYY-MM-DD). Default: from BLIP/Absence")
    args = parser.parse_args()
    if args.expected_hours != EXPECTED_HOURS_PER_DAY:
        print(f"  Using expected hours: {args.expected_hours} (Var = worked - {args.expected_hours})")

    in_path = os.path.abspath(args.input)
    out_path = os.path.abspath(args.output or args.input)

    if not os.path.isfile(in_path):
        print(f"Input file not found: {in_path}", file=__import__("sys").stderr)
        return 1

    # Read sheets
    df_employees = pd.read_excel(in_path, sheet_name="Employees", engine="openpyxl")
    try:
        df_blip = pd.read_excel(in_path, sheet_name="BLIP", engine="openpyxl")
    except Exception:
        df_blip = pd.DataFrame()
    try:
        df_absence = pd.read_excel(in_path, sheet_name="Absence", engine="openpyxl")
    except Exception:
        df_absence = pd.DataFrame()

    # Variance: individual + roll-ups. WFH allowed = 1/week from employee's first absence entry to today (or override)
    df_var = build_variance_individual_df(df_employees, df_blip, df_absence=df_absence, wfh_allowance_override=args.wfh_allowance)
    by_dept = build_variance_rollup(df_var, "Team", "Department") if not df_var.empty else pd.DataFrame()
    by_country = build_variance_rollup(df_var, "Country", "Country") if not df_var.empty else pd.DataFrame()
    by_group = build_variance_rollup(df_var, "Group", "Group") if not df_var.empty else pd.DataFrame()

    # Hours view data (employees and lookup use First+Last)
    dates_list, employees_list, hours_lookup = build_hours_view_data(df_blip, expected_hours=args.expected_hours)

    # Time by Employee: continuous date range (from args or BLIP/Absence)
    time_sheet_start, time_sheet_end = _get_time_sheet_date_range(
        df_blip, df_absence, args.time_from, args.time_to
    )
    time_sheet_dates = pd.date_range(time_sheet_start, time_sheet_end, freq="D").tolist()
    years_in_range = set()
    for d in time_sheet_dates:
        t = pd.Timestamp(d)
        years_in_range.add(t.year)
    national_holidays = set()
    for y in years_in_range:
        national_holidays.update(_uk_holidays_for_year(y))
    absence_lookup = _build_absence_daily_lookup(df_absence)

    # Roll-ups use same period as Time by Employee sheet
    dates_for_rollup = time_sheet_dates if time_sheet_dates else dates_list
    time_by_dept = build_time_rollup(df_employees, hours_lookup, dates_for_rollup, employees_list, "Team", "Department")
    time_by_country = build_time_rollup(df_employees, hours_lookup, dates_for_rollup, employees_list, "Country", "Country")
    time_by_group = build_time_rollup(df_employees, hours_lookup, dates_for_rollup, employees_list, "Group", "Group")

    # New workbook with exactly 8 sheets (no Absence/BLIP/Employees)
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Leave by Employee (cols: ... Taken, % YTD, Remaining, ... WFH Variance)
    ws1 = wb.create_sheet("Leave by Employee", 0)
    _write_variance_sheet(ws1, df_var, "Leave by Employee", variance_col_indices=[8, 11])

    # 2. Time by Employee: continuous calendar, status rules, weekly/monthly/YTD subtotals
    ws2 = wb.create_sheet("Time by Employee", 1)
    ws2["A1"] = "Time by Employee"
    if employees_list and time_sheet_dates:
        col = 2
        emp_col_start = {}
        COLS_PER_EMP = 5  # In, Out, Break, Var, gap
        for emp in employees_list:
            emp_col_start[emp] = col
            ws2.cell(row=2, column=col, value=emp)
            for sub in ["In", "Out", "Break", "Var"]:
                ws2.cell(row=3, column=col, value=sub)
                col += 1
            col += 1  # gap

        # Precompute var value per (emp, date) for subtotals: None=exclude, 0=annual leave, float=BLIP
        def _var_value_for_subtotal(emp, dt):
            ts = pd.Timestamp(dt).normalize()
            if ts.weekday() >= 5 or ts in national_holidays:
                return None
            cats = absence_lookup.get((emp, ts), set())
            if "Medical + Sickness" in cats:
                return None
            if "Annual Leave" in cats:
                return 0.0
            rec = hours_lookup.get((emp, ts), {})
            return float(rec.get("var_hrs", 0) or 0)

        def _cell_value_for_row(emp, dt):
            ts = pd.Timestamp(dt).normalize()
            if ts.weekday() >= 5 or ts in national_holidays:
                return "", "", "", None
            cats = absence_lookup.get((emp, ts), set())
            if "Medical + Sickness" in cats:
                return "no show", "no show", "", None
            if "Annual Leave" in cats:
                return "", "", "", 0.0
            rec = hours_lookup.get((emp, ts), {})
            v = float(rec.get("var_hrs", 0) or 0)
            return (
                rec.get("in", ""),
                rec.get("out", ""),
                _break_hrs_to_hhmm(rec.get("break_hrs", 0)),
                v,
            )

        current_row = 4
        for di, dt in enumerate(time_sheet_dates):
            ts = pd.Timestamp(dt).normalize()
            ws2.cell(row=current_row, column=1, value=ts.strftime("%Y-%m-%d"))
            for emp in employees_list:
                c0 = emp_col_start[emp]
                in_val, out_val, break_val, var_val = _cell_value_for_row(emp, ts)
                ws2.cell(row=current_row, column=c0, value=in_val)
                ws2.cell(row=current_row, column=c0 + 1, value=out_val)
                ws2.cell(row=current_row, column=c0 + 2, value=break_val)
                if var_val is not None:
                    cell = ws2.cell(row=current_row, column=c0 + 3, value=round(var_val, 2))
                    fill = _fill_for_variance(var_val)
                    if fill:
                        cell.fill = fill
            current_row += 1

            # Weekly cumulative (after each Sunday)
            if ts.weekday() == 6:
                week_end = ts
                week_start = week_end - pd.Timedelta(days=6)
                ws2.cell(row=current_row, column=1, value=f"Weekly Cumulative (to {week_end.strftime('%Y-%m-%d')})")
                for emp in employees_list:
                    total = 0.0
                    for d in time_sheet_dates:
                        t = pd.Timestamp(d).normalize()
                        if week_start <= t <= week_end:
                            v = _var_value_for_subtotal(emp, t)
                            if v is not None:
                                total += v
                    c0 = emp_col_start[emp] + 3
                    cell = ws2.cell(row=current_row, column=c0, value=round(total, 2))
                    fill = _fill_for_variance(total)
                    if fill:
                        cell.fill = fill
                current_row += 1

            # Monthly cumulative (after last day of month)
            last_of_month = (ts + pd.offsets.MonthEnd(0)).normalize()
            if ts == last_of_month:
                month_name = ts.strftime("%B")
                ws2.cell(row=current_row, column=1, value=f"{month_name} Cumulative")
                for emp in employees_list:
                    total = 0.0
                    for d in time_sheet_dates:
                        t = pd.Timestamp(d).normalize()
                        if t.month == ts.month and t.year == ts.year:
                            v = _var_value_for_subtotal(emp, t)
                            if v is not None:
                                total += v
                    c0 = emp_col_start[emp] + 3
                    cell = ws2.cell(row=current_row, column=c0, value=round(total, 2))
                    fill = _fill_for_variance(total)
                    if fill:
                        cell.fill = fill
                current_row += 1

        # YTD / Period cumulative (final row)
        ws2.cell(row=current_row, column=1, value="YTD Cumulative Variation")
        for emp in employees_list:
            total = 0.0
            for d in time_sheet_dates:
                v = _var_value_for_subtotal(emp, d)
                if v is not None:
                    total += v
            c0 = emp_col_start[emp] + 3
            cell = ws2.cell(row=current_row, column=c0, value=round(total, 2))
            fill = _fill_for_variance(total)
            if fill:
                cell.fill = fill
    else:
        ws2["A2"] = "No BLIP data or no employees; or no date range for Time sheet."

    # 3. Leave by Department (cols: Name, Headcount, Allowance, Taken, % YTD, Remaining, ... WFH Variance)
    ws3 = wb.create_sheet("Leave by Department", 2)
    _write_variance_sheet(ws3, by_dept, "Leave by Department", variance_col_indices=[6, 9])

    # 4. Time by Department
    ws4 = wb.create_sheet("Time by Department", 3)
    _write_time_rollup_sheet(ws4, time_by_dept, "Time by Department")

    # 5. Leave by Country
    ws5 = wb.create_sheet("Leave by Country", 4)
    _write_variance_sheet(ws5, by_country, "Leave by Country", variance_col_indices=[6, 9])

    # 6. Time by Country
    ws6 = wb.create_sheet("Time by Country", 5)
    _write_time_rollup_sheet(ws6, time_by_country, "Time by Country")

    # 7. Leave by Group
    ws7 = wb.create_sheet("Leave by Group", 6)
    _write_variance_sheet(ws7, by_group, "Leave by Group", variance_col_indices=[6, 9])

    # 8. Time by Group
    ws8 = wb.create_sheet("Time by Group", 7)
    _write_time_rollup_sheet(ws8, time_by_group, "Time by Group")

    # 9. Time Summary: Current Week hours, %, MTD, %, YTD, % (separate report)
    ws9 = wb.create_sheet("Time Summary", 8)
    ws9["A1"] = "Time Summary"
    summary_rows = _build_time_summary_rows(
        employees_list, time_sheet_dates, hours_lookup, absence_lookup, national_holidays, expected_hours_per_day=args.expected_hours
    )
    if summary_rows:
        headers = ["Employee", "Current Week hours", "%", "MTD", "%", "YTD", "%"]
        for c, h in enumerate(headers, 1):
            cell = ws9.cell(row=2, column=c, value=h)
            cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        for ri, row in enumerate(summary_rows, 3):
            ws9.cell(row=ri, column=1, value=row["Employee"])
            ws9.cell(row=ri, column=2, value=row["Current Week hours"])
            pct_cw = row["Current Week %"]
            ws9.cell(row=ri, column=3, value=f"{pct_cw}%" if pct_cw is not None else "")
            if pct_cw is not None and pct_cw < 0:
                ws9.cell(row=ri, column=3).font = Font(color="C62828")
            ws9.cell(row=ri, column=4, value=row["MTD"])
            pct_mtd = row["MTD %"]
            ws9.cell(row=ri, column=5, value=f"{pct_mtd}%" if pct_mtd is not None else "")
            if pct_mtd is not None and pct_mtd < 0:
                ws9.cell(row=ri, column=5).font = Font(color="C62828")
            ws9.cell(row=ri, column=6, value=row["YTD"])
            pct_ytd = row["YTD %"]
            ws9.cell(row=ri, column=7, value=f"{pct_ytd}%" if pct_ytd is not None else "")
            if pct_ytd is not None and pct_ytd < 0:
                ws9.cell(row=ri, column=7).font = Font(color="C62828")
    else:
        ws9["A2"] = "No data (need BLIP and date range for Time sheet)."

    wb.save(out_path)
    print(f"Wrote 9-sheet workbook: {out_path}")
    print(f"  1. Leave by Employee: {len(df_var)} rows")
    print(f"  2. Time by Employee: {len(time_sheet_dates)} dates, {len(employees_list)} employees (with column gap)")
    print(f"  3. Leave by Department: {len(by_dept)} rows")
    print(f"  4. Time by Department: {len(time_by_dept)} rows")
    print(f"  5. Leave by Country: {len(by_country)} rows")
    print(f"  6. Time by Country: {len(time_by_country)} rows")
    print(f"  7. Leave by Group: {len(by_group)} rows")
    print(f"  8. Time by Group: {len(time_by_group)} rows")
    print(f"  9. Time Summary: {len(summary_rows)} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
